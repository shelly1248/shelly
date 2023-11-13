import openpyxl # xlrd = xls 만 가능 

workbook = openpyxl.load_workbook('C:\\sqlite\\mysql\\code\\excel\\새 폴더\\singer.xlsx')
wsheetList = workbook.sheetnames

for wsName in wsheetList:
    worksheet = workbook[wsName]
    print(f'워크시트 이름 : {wsName}')
    for row in range(1, worksheet.max_row+1):
        for col in range(1, worksheet.max_column+1):
            print('%s' % (worksheet.cell(row=row, column=col).value), end='\t')
        print()
    print()