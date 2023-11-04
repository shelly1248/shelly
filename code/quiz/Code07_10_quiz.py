import openpyxl
from copy import copy # 복사 기능

workbook = openpyxl.load_workbook('C:\\sqlite\\mysql\\code\\excel\\새 폴더\\singer.xlsx')
wsheetList = workbook.sheetnames
# print(wsheetList)

outWorkbook = openpyxl.Workbook()
# print(outWorkbook)
outWorkbook.remove(outWorkbook['Sheet']) # 기본으로 생성된 시트를 일단 제거 / 빈워크북 생성(출력용)
# 이름의 기본워크시트가 생김

for wsName in wsheetList:
    worksheet = workbook[wsName]
    # print(wsName)
    outSheet = outWorkbook.create_sheet('New_' + wsName)
    # print(outSheet)
    totalRow = 0
    for row in range(1, worksheet.max_row+1): # +1하는이유 0부터 시작하니까
        # print(row)
        if row != 1 :
            if int(worksheet.cell(row = row , column = 5).value) <=165:
                continue
        totalRow += 1

        outSheet.row_dimensions[row].height = worksheet.row_dimensions[row].height # 행 높이
        
        # print(worksheet.row_dimensions[row].height)
        for col in range(1, worksheet.max_column+1):
            outSheet.column_dimensions[chr(ord('A')+col-1)].width \
                = worksheet.column_dimensions[chr(ord('A')+col-1)].width
#             # [chr(ord('A')+col-1)] A 숫자(65), col = 1 로 시작 col 1은 빼줘야함
            inCell = worksheet.cell(row=row, column=col)
            # #print(inCell)
            outCell = outSheet.cell(row=row, column=col, value= inCell.value)
            # #print(outCell)
            
            if inCell.has_style:
                outCell.font = copy(inCell.font)
                outCell.border = copy(inCell.border)
                outCell.fill = copy(inCell.fill)
                outCell.number_format = copy(inCell.number_format)
                outCell.alignment = copy(inCell.alignment)


outWorkbook.save('\\sqlite\\mysql\\code\\excel\\새 폴더\\singer_copy2.xlsx')
print('save')