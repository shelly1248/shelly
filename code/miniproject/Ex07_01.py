import csv
import glob
import os
import openpyxl

# 전역 변수부
file_list = glob.glob(os.path.join('C:\\sqlite\\mysql\\code\\miniproject\\', '*.xlsx'))
# print(file_list)
firstYN = True # 프로그램의 전역적으로 영향을 미치는 함수

# # 메인 코드
for input_file in file_list:
    workbook = openpyxl.load_workbook(input_file)  
    # print(workbook)
    wsheetList = workbook.sheetnames
    # print(wsheetList)
    for wsName in wsheetList:
        # print(wsName)
        worksheet = workbook[wsName]
        print(worksheet)
        with open('C:\\sqlite\\mysql\\code\\miniproject\\강원인구통합.csv','a', newline='') as outFp: # 'a' = append 있으면 생성X 없으면 생성
            csvWriter = csv.writer(outFp)
#             # =================================================================================== 헤더
#             if firstYN == True:
#                 header_list = []

#                 for col in range(1, worksheet.max_column + 1):
#                     print(col)
#                     header_list.append(worksheet.cell(row=1, column=col).value)
#                 csvWriter.writerow(header_list)
#                 firstYN = False
#             # ====================================================================================
#             for row in range(2, worksheet.max_row + 1):
#                 row_list = []
#                 print(worksheet.max_column)
#                 for col in range(1, worksheet.max_column + 1):
#                     row_list.append(worksheet.cell(row=row, column=col).value)
#                 csvWriter.writerow(row_list)
# print('save')